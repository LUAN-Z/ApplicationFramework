#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Log TXT 排序和 Excel 导出核心逻辑。"""

import os
import re
from dataclasses import dataclass
from pathlib import Path

try:
    from .dependencies import add_plugin_dependency_paths
except ImportError:
    from dependencies import add_plugin_dependency_paths

add_plugin_dependency_paths()

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

CONFIG_MARKER = "<SPEC Max>:"
DEFAULT_HEADER_LINE_COUNT = 38


@dataclass(frozen=True)
class LogRow:
    original_line: str
    batch: int
    port: int
    device_id: int
    timestamp: int
    original_index: int


def try_to_number(cell):
    """尝试将单元格字符串转换为 int 或 float。"""
    text = cell.strip()
    if not text:
        return ""

    if re.fullmatch(r"[+-]?\d+", text):
        try:
            return int(text)
        except ValueError:
            pass

    try:
        return float(text)
    except ValueError:
        return text


def process_txt_line(line, split_spaces=True):
    """处理单行 TXT：有 TAB 按列拆分（split_spaces=True 无 TAB 按空白拆列）。"""
    stripped = line.strip()
    if not stripped:
        return []
    if "\t" in stripped:
        return [try_to_number(cell) for cell in stripped.split("\t")]
    if split_spaces:
        return [try_to_number(cell) for cell in stripped.split()]
    return [try_to_number(stripped)]


def find_keyword(file_name, key_words):
    """返回文件名中命中的第一个关键字，未命中则返回 None。"""
    for key in key_words:
        if key in file_name:
            return key
    return None


def collect_txt_files(input_paths, recursive=False):
    """从文件或目录收集 TXT 文件路径。返回（去重排序后的路径列表, 缺失路径列表）。"""
    txt_files = []
    missing_paths = []

    for raw_path in input_paths:
        path = Path(raw_path)
        if not path.exists():
            missing_paths.append(raw_path)
            continue
        if path.is_file():
            if path.suffix.lower() == ".txt":
                txt_files.append(path)
            continue

        pattern = "**/*.txt" if recursive else "*.txt"
        txt_files.extend(path.glob(pattern))

    unique_sorted = sorted({p.resolve() for p in txt_files})
    return unique_sorted, missing_paths


def extract_device_id(device_str):
    """从设备字段中提取数字编号，未命中时返回 0。"""
    match = re.search(r"\d+", device_str)
    return int(match.group()) if match else 0


def sort_log_data(lines):
    """核心排序逻辑：解析数据行、按'端口号重复'识别批次、按 (Batch, Port, Device, Time) 排序。"""
    data_rows = []
    current_batch = 0
    seen_ports_in_current_batch = set()

    for idx, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue

        parts = line.split()

        if len(parts) < 7:
            continue

        try:
            port = int(parts[4])

            if port in seen_ports_in_current_batch:
                current_batch += 1
                seen_ports_in_current_batch.clear()

            seen_ports_in_current_batch.add(port)
            data_rows.append(
                LogRow(
                    original_line=line,
                    batch=current_batch,
                    port=port,
                    device_id=extract_device_id(parts[3]),
                    timestamp=int(parts[6]),
                    original_index=idx,
                )
            )

        except ValueError:
            continue

    sorted_data = sorted(
        data_rows,
        key=lambda row: (row.batch, row.port, row.device_id, row.timestamp, row.original_index),
    )
    return sorted_data, current_batch + 1


def find_last_config_line_idx(lines):
    """查找最后一行配置的索引，未找到返回 -1。"""
    for idx in range(len(lines) - 1, -1, -1):
        line = lines[idx]
        if line.startswith(CONFIG_MARKER):
            return idx
    return -1


def split_header_and_data(all_lines):
    """按配置段位置切分头部与数据区域。"""
    header_count = find_last_config_line_idx(all_lines) + 1
    if header_count <= 0:
        header_count = DEFAULT_HEADER_LINE_COUNT

    header_end = min(header_count, len(all_lines))
    return all_lines[:header_end], all_lines[header_end:]


def write_lines(file_obj, lines):
    """逐行写入，自动补齐换行符。"""
    for line in lines:
        if line.endswith("\n"):
            file_obj.write(line)
        else:
            file_obj.write(line + "\n")


DEFAULT_KEYWORDS = ("NTNV", "NTHV", "NTLV", "HTNV", "HTHV", "HTLV")
INVALID_SHEET_CHARS = re.compile(r'[:\\/?*\[\]]')


def build_sorted_content(input_file, encoding="utf-8", keep_header=True):
    try:
        all_lines = input_file.read_text(encoding=encoding).splitlines(keepends=True)
    except OSError as exc:
        raise OSError(f"读取文件失败 {input_file}: {exc}") from exc

    header_lines, data_lines = split_header_and_data(all_lines)
    sorted_data, batch_count = sort_log_data(data_lines)
    sorted_lines = [item.original_line for item in sorted_data]

    final_lines = []
    if keep_header:
        final_lines.extend(line.rstrip("\n") for line in header_lines)
    final_lines.extend(sorted_lines)

    # log汇总只保留 CONFIG_MARKER 行 + 往上3行
    summary_header_lines = []
    if keep_header and header_lines:
        start_idx = max(0, len(header_lines) - 4)
        summary_header_lines = [line.rstrip("\n") for line in header_lines[start_idx:]]

    return {
        "path": input_file,
        "lines": final_lines,
        "batch_count": batch_count,
        "sorted_row_count": len(sorted_data),
        "header_lines": [line.rstrip("\n") for line in header_lines] if keep_header else [],
        "sorted_lines": sorted_lines,
        "summary_header_lines": summary_header_lines,
    }


def save_sorted_txt(processed_file, sorted_dir=None, encoding="utf-8"):
    input_file = processed_file["path"]

    if sorted_dir:
        output_dir = Path(sorted_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / f"{input_file.stem}_sorted{input_file.suffix}"
    else:
        output_path = input_file.with_name(f"{input_file.stem}_sorted{input_file.suffix}")

    with output_path.open("w", encoding=encoding, newline="\n") as file_obj:
        write_lines(file_obj, processed_file["lines"])

    return output_path


def replace_all_placeholder(line, sheet_name):
    """把行内的占位符 ALL 替换为 sheet 名。仅替换独立 ALL 词（词边界匹配）。"""
    return re.sub(r"\bALL\b", sheet_name, line)


def sanitize_sheet_name(name, fallback="Sheet"):
    cleaned = INVALID_SHEET_CHARS.sub("_", name).strip()
    if not cleaned:
        cleaned = fallback
    return cleaned[:31]


def make_unique_sheet_name(workbook, base_name):
    candidate = sanitize_sheet_name(base_name)
    if candidate not in workbook.sheetnames:
        return candidate

    index = 2
    while True:
        suffix = f"_{index}"
        trimmed = candidate[: 31 - len(suffix)]
        unique_name = f"{trimmed}{suffix}"
        if unique_name not in workbook.sheetnames:
            return unique_name
        index += 1


def build_sheet_plan(processed_files, key_words, import_unmatched=False):
    ordered_keywords = list(dict.fromkeys(key_words))
    grouped_files = {key: [] for key in ordered_keywords}
    skipped_files = []
    sheet_plan = []

    for item in processed_files:
        matched_key = find_keyword(item["path"].name, ordered_keywords)
        if matched_key:
            grouped_files[matched_key].append(item)
        elif import_unmatched:
            sheet_plan.append((item["path"].stem, [item]))
        else:
            skipped_files.append(item["path"].name)

    for key in ordered_keywords:
        files = grouped_files[key]
        if files:
            sheet_plan.append((key, files))

    return sheet_plan, skipped_files


def export_to_excel(processed_files, excel_file_path, key_words=DEFAULT_KEYWORDS, append_existing=False, import_unmatched=False):
    sheet_plan, skipped_files = build_sheet_plan(
        processed_files=processed_files,
        key_words=key_words,
        import_unmatched=import_unmatched,
    )

    if not sheet_plan:
        print("警告：没有文件命中关键字，未生成 Excel")
        return False, skipped_files

    try:
        excel_path = Path(excel_file_path)
        if append_existing and excel_path.exists():
            workbook = load_workbook(excel_path)
        else:
            workbook = Workbook()
            default_sheet = workbook.active
            if default_sheet.max_row == 1 and default_sheet.max_column == 1 and default_sheet["A1"].value is None:
                workbook.remove(default_sheet)

        data_font = Font(name="Microsoft YaHei")

        for sheet_base_name, files in sheet_plan:
            sheet_name = make_unique_sheet_name(workbook, sheet_base_name)
            worksheet = workbook.create_sheet(title=sheet_name)
            row = 1
            print(f"[OK] 新建 sheet: {sheet_name}，来源组: {sheet_base_name}，文件数: {len(files)}")
            for item in files:
                print(f"正在写入文件：{item['path'].name}")
                for line in item["lines"]:
                    line = replace_all_placeholder(line, sheet_name)
                    row_data = process_txt_line(line)
                    if not row_data:
                        continue
                    for col, value in enumerate(row_data, start=1):
                        cell = worksheet.cell(row=row, column=col, value=value)
                        cell.font = data_font
                    row += 1
                row += 1

        # --- 创建 log汇总 sheet ---
        summary_name = make_unique_sheet_name(workbook, "log汇总")
        summary_ws = workbook.create_sheet(title=summary_name)
        summary_row = 1
        print(f"[OK] 新建 sheet: {summary_name}，汇总所有文件数据")

        for _sheet_base_name, files in sheet_plan:
            for item in files:
                for line in item.get("summary_header_lines", []):
                    line = replace_all_placeholder(line, _sheet_base_name)
                    row_data = process_txt_line(line)
                    if not row_data:
                        continue
                    for col, value in enumerate(row_data, start=1):
                        cell = summary_ws.cell(row=summary_row, column=col, value=value)
                        cell.font = data_font
                    summary_row += 1

                for line in item.get("sorted_lines", []):
                    line = replace_all_placeholder(line, _sheet_base_name)
                    row_data = process_txt_line(line)
                    if not row_data:
                        continue
                    for col, value in enumerate(row_data, start=1):
                        cell = summary_ws.cell(row=summary_row, column=col, value=value)
                        cell.font = data_font
                    summary_row += 1

                summary_row += 1  # 文件之间空一行

        workbook.save(excel_path)
    except Exception as exc:
        print(f"创建或写入 Excel 失败：{exc}")
        return False, skipped_files

    print(f"Excel 文件保存至：{os.path.abspath(excel_file_path)}")
    return True, skipped_files


def resolve_output_path(output, txt_files, default_name="result.xlsx"):
    """确定输出 Excel 的路径。

    用户显式指定时原样使用；否则把输出文件放到日志所在目录，
    目录取所有日志文件的共同父目录（并集目录）。
    """
    if output:
        return Path(output)

    if not txt_files:
        return Path(default_name)

    parent_dirs = {p.parent for p in txt_files}
    if len(parent_dirs) == 1:
        base_dir = next(iter(parent_dirs))
    else:
        base_dir = Path(os.path.commonpath([str(p) for p in parent_dirs]))

    return base_dir / default_name


